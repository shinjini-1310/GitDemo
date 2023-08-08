#THIS IS THE ALTERED VERSION OF THE CODE DOWNLOADED/CLONED FROM ORIGINAL CODE IN REMOVE GITHUB REPO
#2nd time alteration done by user GitA after downloading the altered code uploaded by GitDemo to remote repo
#USER NO 2 CHECKED OUT THE LATEST CODE FROM MASTER BRANCH INTO A NEW DEV BRANCH, MODIFIED CODE AND MERGED IT TO MASTER
#ORIGINAL CODE CREATOR PULLED LATEST DEV BRANCH GITHUB CODE INTO THIS LOCAL DEV BRANCH AMD MODIFIED IT (ADDED PALINDROME CODE)

import openpyxl as op

book = op.load_workbook("C:\\Users\\SHINJINI\\Imp_Docs\\Selenium Course\\pythonDemo.xlsx")
sheet = book.active
data_dict = {}
cell = sheet.cell(row=1,column=2)

for row in range(1,sheet.max_row+1):
    if sheet.cell(row=row,column=1).value == "Testcase2":
        for column in range(2,sheet.max_column+1):
            #print(sheet.cell(row=row,column=column).value)
            data_dict[sheet.cell(row=1,column=column).value]=sheet.cell(row=row,column=column).value

# function which return reverse of a string
def isPalindrome(s):
    return s == s[::-1]
# Driver code
s = "malayalam"
ans = isPalindrome(s)

if ans: print("Yes")
else: print("No")
            
print("THIS IS THE ALTERED VERSION OF THE CODE DOWNLOADED/CLONED FROM ORIGINAL CODE IN REMOVE GITHUB REPO")
print("2ND TIME ALTERATION DONE BY USER GITA AFTER DOWNLOADING THE ALTERED CODE UPLOADED BY GITDEMO TO REMOTE REPO")
print("USER NO 2 CHECKED OUT THE LATEST CODE FROM MASTER BRANCH INTO A NEW DEV BRANCH, MODIFIED CODE AND MERGED IT TO MASTER")
print("ORIGINAL CODE CREATOR PULLED LATEST DEV BRANCH GITHUB CODE INTO THIS LOCAL DEV BRANCH AMD MODIFIED IT (ADDED PALINDROME CODE)")